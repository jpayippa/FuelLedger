import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font


def _bold_row(ws, row_idx):
    for cell in ws[row_idx]:
        cell.font = Font(bold=True)


def _money_column(ws, col_letter, min_row=2):
    for row in ws.iter_rows(min_row=min_row, min_col=ws[col_letter + "1"].column,
                             max_col=ws[col_letter + "1"].column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0.00"


def build_workbook(fuel_logs, maintenance_logs, odometer_logs, monthly_totals, yearly_totals, per_vehicle_summary):
    wb = Workbook()

    ws = wb.active
    ws.title = "Fuel Logs"
    ws.append(["Date", "Vehicle ID", "Station", "Amount", "Volume", "Unit", "Price/Unit", "Odometer"])
    _bold_row(ws, 1)
    for r in fuel_logs:
        ws.append([r["date"], r["vehicle_id"], r.get("station") or "", r["amount_cents"] / 100,
                   r.get("volume"), r.get("volume_unit") or "", r.get("price_per_unit"), r.get("odometer")])
    for col, width in zip("ABCDEFGH", (12, 10, 18, 10, 9, 6, 11, 10)):
        ws.column_dimensions[col].width = width
    _money_column(ws, "D")

    ws2 = wb.create_sheet("Maintenance Logs")
    ws2.append(["Date", "Vehicle ID", "Shop", "Category", "Amount", "Odometer", "Notes"])
    _bold_row(ws2, 1)
    for r in maintenance_logs:
        category = r.get("category_other") or r.get("category")
        ws2.append([r["date"], r["vehicle_id"], r.get("shop") or "", category,
                    r["amount_cents"] / 100, r.get("odometer"), r.get("notes") or ""])
    for col, width in zip("ABCDEFG", (12, 10, 18, 18, 10, 10, 24)):
        ws2.column_dimensions[col].width = width
    _money_column(ws2, "E")

    ws3 = wb.create_sheet("Odometer Logs")
    ws3.append(["Date", "Vehicle ID", "Odometer", "Note"])
    _bold_row(ws3, 1)
    for r in odometer_logs:
        ws3.append([r["date"], r["vehicle_id"], r["odometer"], r.get("note") or ""])
    for col, width in zip("ABCD", (12, 10, 10, 24)):
        ws3.column_dimensions[col].width = width

    ws4 = wb.create_sheet("Monthly Summary")
    ws4.append(["Month", "Fuel Total", "Fill-ups", "Avg Price/Unit"])
    _bold_row(ws4, 1)
    for m in monthly_totals:
        ws4.append([m["period"], m["total_cents"] / 100, m["count"],
                    round(m["avg_price_per_unit"], 3) if m.get("avg_price_per_unit") else None])
    for col, width in zip("ABCD", (12, 12, 10, 14)):
        ws4.column_dimensions[col].width = width
    _money_column(ws4, "B")

    ws5 = wb.create_sheet("Yearly Summary")
    ws5.append(["Year", "Fuel Total", "Fill-ups"])
    _bold_row(ws5, 1)
    for y in yearly_totals:
        ws5.append([y["year"], y["total_cents"] / 100, y["count"]])
    for col, width in zip("ABC", (10, 12, 10)):
        ws5.column_dimensions[col].width = width
    _money_column(ws5, "B")

    ws6 = wb.create_sheet("Summary")
    ws6.append(["Vehicle", "Fuel Total", "Fill-ups", "Avg Fill", "Maintenance Total",
                "Maintenance Count", "Combined Total", "Cost/km"])
    _bold_row(ws6, 1)
    for v in per_vehicle_summary:
        ws6.append([
            v["vehicle"], v["fuel_total_cents"] / 100, v["fuel_count"], v["avg_fill_cents"] / 100,
            v["maintenance_total_cents"] / 100, v["maintenance_count"], v["combined_total_cents"] / 100,
            v["cost_per_km"] if v["cost_per_km"] is not None else "N/A",
        ])
    for col, width in zip("ABCDEFGH", (18, 12, 10, 10, 16, 12, 14, 10)):
        ws6.column_dimensions[col].width = width
    for col in ("B", "D", "E", "G"):
        _money_column(ws6, col)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_csv(fuel_logs):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date", "Vehicle ID", "Station", "Amount", "Volume", "Unit", "Price/Unit", "Odometer"])
    total_cents = 0
    for r in fuel_logs:
        total_cents += r["amount_cents"]
        writer.writerow([
            r["date"], r["vehicle_id"], r.get("station") or "", f'{r["amount_cents"] / 100:.2f}',
            r.get("volume") or "", r.get("volume_unit") or "", r.get("price_per_unit") or "",
            r.get("odometer") or "",
        ])
    writer.writerow([])
    writer.writerow(["Total", "", "", f"{total_cents / 100:.2f}"])
    return io.BytesIO(buf.getvalue().encode("utf-8"))
